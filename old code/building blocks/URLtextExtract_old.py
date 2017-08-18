#call in necessary packages that code requires to operate
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
##################################################
##################################################

#define number of loop iterations (i.e. how many links to analyze)
num = 2

#load in the excel workbook
wb = openpyxl.load_workbook('xml_searchResults.xlsx')

#load working directory for file saving of files containing the URL text
os.chdir('/Users/PMcG/Dropbox (ASU)/AAB_files/Pat-files/WCP/code/outputFiles/')

#retrieve the sheet names
wb.get_sheet_names()

#use the called names to define objects for each search engine
google = wb.get_sheet_by_name('google')
bing = wb.get_sheet_by_name('bing') 
yahoo = wb.get_sheet_by_name('yahoo')

#loop through each search engine's page and call the individual URL text
#for x in range (1,num):
#    print google.cell(row = num, column = 1).value 
#    print bing.cell(row = num, column = 1).value 
#    print yahoo.cell(row = num, column = 1).value 

##################################################
##################################################
                                    ###GOOGLE###
##################################################
##################################################

#define parameters for url text removal
prefix1 = ("/url?q=")
prefix2 = ("&sa=U&ved=0ahUKEwitrfTds7rNAhUC2oMKHZhtCtQQFggcMAI&usg=AFQjCNH6xwqVs8ZO4wckbL5qQU_XEZ65Hw'")

for x in range (1,num):

    #remove excell text from each URL
    temp_dat = str(google.cell(row = x, column = 1).value)   
    urlReady = temp_dat.replace(prefix1, "", 1)
    urlReady = urlReady.replace(prefix2, "", 1)

    #verify URL is correct and ready to be accessed via web server
    print(urlReady)

    #access web server and retrieve web content
    url = urlReady
    r = requests.get(url)
    soup = BeautifulSoup(r.content)

    #strip HTML scripting and style
    for script in soup(["script", "style"]):
            script.extract()    # rip it out

    #clean up the text
    # get text
    text = soup.get_text()
    # break into lines and remove leading and trailing space on each
    lines = (line.strip() for line in text.splitlines())
    # break multi-headlines into a line each
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    # drop blank lines
    text = '\n'.join(chunk for chunk in chunks if chunk)

    #write contents to file based on the internet search engine and loop count data
    write_text = text.encode('ascii','ignore')

    fileName = "google_"  + str(x) + ".txt"
    f = open(fileName, 'a');

    f.write(write_text)
    f.close()

    print(text)

##################################################
##################################################
                                    ###BING###
##################################################
##################################################

for x in range (1,num):

    #remove excell text from each URL
    temp_dat = str(bing.cell(row = x, column = 1).value)   
    urlReady = temp_dat.replace(prefix1, "", 1)
    urlReady = urlReady.replace(prefix2, "", 1)

    #verify URL is correct and ready to be accessed via web server
    print(urlReady)

    #access web server and retrieve web content
    url = urlReady
    r = requests.get(url)
    soup = BeautifulSoup(r.content)

    #strip HTML scripting and style
    for script in soup(["script", "style"]):
            script.extract()    # rip it out

    #clean up the text
    # get text
    text = soup.get_text()
    # break into lines and remove leading and trailing space on each
    lines = (line.strip() for line in text.splitlines())
    # break multi-headlines into a line each
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    # drop blank lines
    text = '\n'.join(chunk for chunk in chunks if chunk)

    #write contents to file
    write_text = text.encode('ascii','ignore')

    fileName = "bing_"  + str(x) + ".txt"
    f = open(fileName, 'a');

    f.write(write_text)
    f.close()

    print(text)    

##################################################
##################################################
                                    ###YAHOO###
##################################################
##################################################

#define parameters for url text removal
prefix1 = ("ri.search.yahoo.com/_ylt=AwrTHRNQ32lXVzUAKDNXNyoA;_ylu=X3oDMTByb2lvbXVuBGNvbG8DZ3ExBHBvcwMxBHZ0aWQDBHNlYwNzcg--/RV=2/RE=1466585040/RO=10/RU=https%3a%2f%2fen.")
prefix2 = ("RK=0/RS=1xD4kqsgwa5Mj7hF7IHqw2oscKI-")
prefix3 = ("%2f")

for x in range (1,num):

    #remove excell text from each URL
    temp_dat = str(yahoo.cell(row = x, column = 1).value)   
    urlReady = temp_dat.replace(prefix1, "", 1)
    urlReady = urlReady.replace(prefix2, "", 1)
    urlReady = urlReady.replace(prefix3, "", 1)

    #verify URL is correct and ready to be accessed via web server
    print(urlReady)

    #access web server and retrieve web content
    url = urlReady
    r = requests.get(url)
    soup = BeautifulSoup(r.content)

    #strip HTML scripting and style
    for script in soup(["script", "style"]):
            script.extract()    # rip it out

    #clean up the text
    # get text
    text = soup.get_text()
    # break into lines and remove leading and trailing space on each
    lines = (line.strip() for line in text.splitlines())
    # break multi-headlines into a line each
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    # drop blank lines
    text = '\n'.join(chunk for chunk in chunks if chunk)

    #write contents to file based on the internet search engine and loop count data
    write_text = text.encode('ascii','ignore')

    fileName = "yahoo_"  + str(x) + ".txt"
    f = open(fileName, 'a');

    f.write(write_text)
    f.close()

    print(text)
