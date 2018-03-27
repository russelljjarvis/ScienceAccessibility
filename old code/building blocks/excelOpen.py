#import json as simplejson
#from openpyxl import Workbook
import openpyxl

#load in the excel workbook
wb = openpyxl.load_workbook('xml_searchResults.xlsx')

#retrieve the sheet names
print wb.get_sheet_names()

#use the called names to define objects for each search engine
google = wb.get_sheet_by_name('google')
bing = wb.get_sheet_by_name('bing') 
yahoo = wb.get_sheet_by_name('yahoo')

#loop through each search engine's page and call the individual URL text
for x in range (1,10):
    print (x,google.cell(row = x, column = 1).value )
    print (x,bing.cell(row = x, column = 1).value )
    print (x,yahoo.cell(row = x, column = 1).value )


for x in range (1,10):
    temp_url = (x,google.cell(row = x, column = 1).value )
